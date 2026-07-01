"""
ONGC Compliance Report Generator
==================================
Reads raw JSON results produced by compliance_engine.py and generates:
  - reports/Vendor_A_Compliance.xlsx
  - reports/Vendor_B_Compliance.xlsx
  - reports/Vendor_C_Compliance.xlsx
  - reports/Comparison_Summary.xlsx   (side-by-side all vendors)

Each vendor report contains:
  Sheet 1 — "Compliance Matrix"  : full clause-by-clause evaluation
  Sheet 2 — "Summary"            : aggregated statistics + compliance %

Comparison_Summary.xlsx contains:
  Sheet 1 — "Vendor Comparison"  : all vendors side-by-side per clause
  Sheet 2 — "Score Card"         : headline scores for quick decision-making

Usage
-----
    cd ONGC_RAG/
    python compliance_report_generator.py

    # Specific vendor:
    python compliance_report_generator.py --vendor Vendor_A

Place this file at:
    ONGC_RAG/compliance_report_generator.py
"""

import argparse
import json
import sys
from pathlib import Path
from datetime import datetime

# ── Dependency check ──────────────────────────────────────────────────────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
        GradientFill,
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
except ImportError:
    sys.exit("Run: pip install openpyxl")

# =====================================
# PATHS
# =====================================
RESULTS_DIR  = Path("compliance_results")
REPORTS_DIR  = Path("reports")

# =====================================
# STATUS CONFIG
# =====================================
STATUS_COMPLIANT     = "COMPLIANT"
STATUS_NON_COMPLIANT = "NON_COMPLIANT"
STATUS_NOT_FOUND     = "NOT_FOUND"

# Colours (ARGB hex — openpyxl uses no '#')
CLR_HEADER_BG    = "FF1F3864"   # dark navy (ONGC blue)
CLR_HEADER_FONT  = "FFFFFFFF"   # white
CLR_COMPLIANT    = "FFD9EAD3"   # light green
CLR_NON_COMPLY   = "FFFCE8E6"   # light red
CLR_NOT_FOUND    = "FFFFF3E0"   # light amber
CLR_SECTION_BG   = "FFD0E4F5"   # light blue for alternating sections
CLR_SUMMARY_BG   = "FFF3F3F3"   # light grey for summary labels
CLR_OCR_WARN     = "FFFFF176"   # yellow for OCR warning cells

# Status → fill colour mapping
STATUS_FILL = {
    STATUS_COMPLIANT:     CLR_COMPLIANT,
    STATUS_NON_COMPLIANT: CLR_NON_COMPLY,
    STATUS_NOT_FOUND:     CLR_NOT_FOUND,
}

# Status → display text
STATUS_DISPLAY = {
    STATUS_COMPLIANT:     "✓ Compliant",
    STATUS_NON_COMPLIANT: "✗ Non-Compliant",
    STATUS_NOT_FOUND:     "? Not Found",
    "DRY_RUN":            "— Dry Run",
}


# =====================================
# STYLE HELPERS
# =====================================
def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def _font(bold: bool = False, color: str = "FF000000", size: int = 10,
          italic: bool = False) -> Font:
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="top", wrap_text=True)


def _thin_border() -> Border:
    s = Side(style="thin", color="FFB0B0B0")
    return Border(left=s, right=s, top=s, bottom=s)


def _style_header_row(ws, row: int, ncols: int):
    """Apply dark navy header style to an entire row."""
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill   = _fill(CLR_HEADER_BG)
        cell.font   = _font(bold=True, color=CLR_HEADER_FONT, size=10)
        cell.alignment = _center()
        cell.border = _thin_border()


def _autofit_columns(ws, min_w: int = 10, max_w: int = 60):
    """Approximate column width based on max content length."""
    for col_cells in ws.columns:
        length = max(
            (len(str(cell.value or "")) for cell in col_cells),
            default=min_w,
        )
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max(min_w, min(length + 2, max_w))


# =====================================
# VENDOR REPORT BUILDER
# =====================================
def build_vendor_report(vendor: str, results: list[dict]) -> Workbook:
    """
    Build a complete vendor compliance workbook.
    Returns the Workbook object (caller saves it).
    """
    wb = Workbook()

    # ── Sheet 1: Compliance Matrix ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Compliance Matrix"
    ws.freeze_panes = "A3"   # freeze header rows

    # Title row
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value     = f"ONGC Tender Compliance Evaluation — {vendor}"
    title_cell.font      = _font(bold=True, size=13, color=CLR_HEADER_FONT)
    title_cell.fill      = _fill(CLR_HEADER_BG)
    title_cell.alignment = _center()
    ws.row_dimensions[1].height = 28

    # Column headers
    headers = [
        "Clause No.",
        "Requirement",
        "Remarks",
        "Status",
        "Explanation",
        "Evidence Pages",
        "Source Documents",
        "Confidence",
    ]
    for col_i, hdr in enumerate(headers, 1):
        ws.cell(row=2, column=col_i, value=hdr)
    _style_header_row(ws, 2, len(headers))
    ws.row_dimensions[2].height = 20

    # Column widths (set before data for clarity)
    col_widths = [12, 45, 25, 16, 55, 16, 35, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Data rows
    for row_i, r in enumerate(results, 3):
        status       = r.get("status",           STATUS_NOT_FOUND)
        pages        = r.get("evidence_pages",   [])
        sources      = r.get("evidence_sources", [])
        ocr_warn     = r.get("ocr_warning",      False)
        status_label = STATUS_DISPLAY.get(status, status)

        pages_str   = ", ".join(str(p) for p in pages)  if pages   else "—"
        sources_str = "\n".join(sources)                if sources else "—"

        row_data = [
            r.get("clause_no",   ""),
            r.get("requirement", ""),
            r.get("remarks",     ""),
            status_label,
            r.get("explanation", ""),
            pages_str,
            sources_str,
            r.get("confidence",  ""),
        ]

        row_fill_color = STATUS_FILL.get(status, CLR_NOT_FOUND)

        for col_i, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.fill      = _fill(row_fill_color)
            cell.border    = _thin_border()
            cell.alignment = _center() if col_i in (1, 4, 6, 8) else _left()
            cell.font      = _font(size=9)

            # Bold the status cell
            if col_i == 4:
                cell.font = _font(bold=True, size=9)

        # OCR warning highlight on explanation cell
        if ocr_warn:
            exp_cell      = ws.cell(row=row_i, column=5)
            exp_cell.fill = _fill(CLR_OCR_WARN)

        ws.row_dimensions[row_i].height = 42

    # ── Sheet 2: Summary ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    _build_summary_sheet(ws2, vendor, results)

    return wb


def _build_summary_sheet(ws, vendor: str, results: list[dict]):
    """Populate the Summary sheet with statistics and a simple chart data range."""
    total       = len(results)
    compliant   = sum(1 for r in results if r["status"] == STATUS_COMPLIANT)
    nc          = sum(1 for r in results if r["status"] == STATUS_NON_COMPLIANT)
    not_found   = sum(1 for r in results if r["status"] == STATUS_NOT_FOUND)
    pct         = compliant / total if total else 0
    low_ocr     = sum(1 for r in results if r.get("ocr_warning"))

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20

    # Title
    ws.merge_cells("A1:C1")
    ws["A1"].value     = f"Compliance Summary — {vendor}"
    ws["A1"].font      = _font(bold=True, size=13, color=CLR_HEADER_FONT)
    ws["A1"].fill      = _fill(CLR_HEADER_BG)
    ws["A1"].alignment = _center()

    ws.merge_cells("A2:C2")
    ws["A2"].value     = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font      = _font(italic=True, size=9, color="FF606060")
    ws["A2"].alignment = _center()

    # Stats table
    stats = [
        ("", "Count", "Percentage"),
        ("Total Clauses",    total,      "=B5/B5"),
        ("✓ Compliant",      compliant,  f"=B6/B5"),
        ("✗ Non-Compliant",  nc,         f"=B7/B5"),
        ("? Not Found",      not_found,  f"=B8/B5"),
        ("⚠ OCR Warnings",   low_ocr,    f"=B9/B5"),
    ]

    row_map = {0: 3, 1: 5, 2: 6, 3: 7, 4: 8, 5: 9}
    for i, (label, count, pct_formula) in enumerate(stats):
        r = i + 3 if i == 0 else row_map[i]
        ws.cell(row=r, column=1, value=label).font  = _font(bold=(i == 0), size=10)
        ws.cell(row=r, column=2, value=count).font  = _font(size=10)
        if i > 0:
            pct_cell = ws.cell(row=r, column=3, value=pct_formula)
            pct_cell.number_format = "0.0%"
            pct_cell.font = _font(size=10)

    _style_header_row(ws, 3, 3)

    # Colour-code stat rows
    fills = {
        5: CLR_COMPLIANT,
        6: CLR_NON_COMPLY,
        7: CLR_NOT_FOUND,
        9: CLR_OCR_WARN,
    }
    for row_num, fill_clr in fills.items():
        for col in range(1, 4):
            ws.cell(row=row_num, column=col).fill = _fill(fill_clr)

    # Compliance score highlight
    ws.merge_cells("A11:C11")
    score_cell = ws["A11"]
    score_cell.value     = f"Overall Compliance Score: =B6/B5"
    score_cell.value     = f"Overall Compliance Score: {pct:.1%}"
    score_cell.font      = _font(bold=True, size=14)
    score_cell.alignment = _center()
    score_cell.fill      = _fill(CLR_COMPLIANT if pct >= 0.8 else CLR_NON_COMPLY)

    # Non-compliant clause list
    nc_rows = [r for r in results if r["status"] == STATUS_NON_COMPLIANT]
    if nc_rows:
        ws["A13"].value = "Non-Compliant Clauses"
        ws["A13"].font  = _font(bold=True, color="FFCC0000")
        for j, r in enumerate(nc_rows, 14):
            ws.cell(row=j, column=1, value=r["clause_no"])
            ws.cell(row=j, column=2, value=r["requirement"][:80])
            ws.cell(row=j, column=3, value=r.get("explanation", "")[:100])
            for col in range(1, 4):
                ws.cell(row=j, column=col).fill = _fill(CLR_NON_COMPLY)


# =====================================
# COMPARISON WORKBOOK BUILDER
# =====================================
def build_comparison_report(all_results: dict) -> Workbook:
    """
    Build a side-by-side comparison workbook for all vendors.
    all_results = {vendor_name: [result_dict, ...]}
    """
    wb     = Workbook()
    ws     = wb.active
    ws.title = "Vendor Comparison"
    ws.freeze_panes = "D3"

    vendors = sorted(all_results.keys())
    n_v     = len(vendors)

    # Collect all clauses in order
    first_results = list(all_results.values())[0]
    clauses = [(r["clause_no"], r["requirement"]) for r in first_results]

    # Build lookup: {vendor: {clause_no: result}}
    lookup: dict[str, dict] = {
        v: {r["clause_no"]: r for r in results}
        for v, results in all_results.items()
    }

    # ── Title ──────────────────────────────────────────────────────────────────
    total_cols = 3 + n_v * 2   # clause | req | remarks | (status + pages) × vendors
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    ws["A1"].value     = "ONGC Multi-Vendor Compliance Comparison"
    ws["A1"].font      = _font(bold=True, size=14, color=CLR_HEADER_FONT)
    ws["A1"].fill      = _fill(CLR_HEADER_BG)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 28

    # ── Column headers ─────────────────────────────────────────────────────────
    base_headers = ["Clause No.", "Requirement", "Remarks"]
    for col_i, h in enumerate(base_headers, 1):
        ws.cell(row=2, column=col_i, value=h)

    col_offset = len(base_headers) + 1
    for v_i, vendor in enumerate(vendors):
        col_status = col_offset + v_i * 2
        col_pages  = col_offset + v_i * 2 + 1
        ws.cell(row=2, column=col_status, value=f"{vendor}\nStatus")
        ws.cell(row=2, column=col_pages,  value=f"{vendor}\nPages")

    _style_header_row(ws, 2, total_cols)
    ws.row_dimensions[2].height = 30

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 20
    for v_i in range(n_v):
        ws.column_dimensions[get_column_letter(col_offset + v_i * 2)].width     = 16
        ws.column_dimensions[get_column_letter(col_offset + v_i * 2 + 1)].width = 14

    # ── Data rows ──────────────────────────────────────────────────────────────
    for row_i, (clause_no, requirement) in enumerate(clauses, 3):
        ws.cell(row=row_i, column=1, value=clause_no).font   = _font(size=9)
        ws.cell(row=row_i, column=2, value=requirement).font = _font(size=9)
        ws.cell(row=row_i, column=2).alignment = _left()

        remarks = ""
        for vendor in vendors:
            r = lookup[vendor].get(clause_no, {})
            if r.get("remarks"):
                remarks = r["remarks"]
                break
        ws.cell(row=row_i, column=3, value=remarks).font = _font(size=9, italic=True)

        for v_i, vendor in enumerate(vendors):
            r       = lookup[vendor].get(clause_no, {})
            status  = r.get("status", STATUS_NOT_FOUND)
            pages   = r.get("evidence_pages", [])
            pages_s = ", ".join(str(p) for p in pages) if pages else "—"
            label   = STATUS_DISPLAY.get(status, status)
            fill_c  = STATUS_FILL.get(status, CLR_NOT_FOUND)

            col_status = col_offset + v_i * 2
            col_pages  = col_offset + v_i * 2 + 1

            sc = ws.cell(row=row_i, column=col_status, value=label)
            sc.fill      = _fill(fill_c)
            sc.font      = _font(bold=True, size=9)
            sc.alignment = _center()
            sc.border    = _thin_border()

            pc = ws.cell(row=row_i, column=col_pages, value=pages_s)
            pc.fill      = _fill(fill_c)
            pc.font      = _font(size=9)
            pc.alignment = _center()
            pc.border    = _thin_border()

        ws.row_dimensions[row_i].height = 30

    # ── Sheet 2: Score Card ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Score Card")
    _build_scorecard(ws2, all_results, vendors)

    return wb


def _build_scorecard(ws, all_results: dict, vendors: list[str]):
    """One-glance scorecard comparing all vendors."""
    ws.column_dimensions["A"].width = 25
    for i, _ in enumerate(vendors):
        ws.column_dimensions[get_column_letter(i + 2)].width = 18

    # Title
    ws.merge_cells(f"A1:{get_column_letter(len(vendors) + 1)}1")
    ws["A1"].value     = "Vendor Score Card"
    ws["A1"].font      = _font(bold=True, size=14, color=CLR_HEADER_FONT)
    ws["A1"].fill      = _fill(CLR_HEADER_BG)
    ws["A1"].alignment = _center()

    # Header row
    ws.cell(row=2, column=1, value="Metric")
    for i, v in enumerate(vendors, 2):
        ws.cell(row=2, column=i, value=v)
    _style_header_row(ws, 2, len(vendors) + 1)

    metrics = [
        ("Total Clauses",       lambda r: len(r)),
        ("✓ Compliant",         lambda r: sum(1 for x in r if x["status"] == STATUS_COMPLIANT)),
        ("✗ Non-Compliant",     lambda r: sum(1 for x in r if x["status"] == STATUS_NON_COMPLIANT)),
        ("? Not Found",         lambda r: sum(1 for x in r if x["status"] == STATUS_NOT_FOUND)),
        ("⚠ OCR Warnings",      lambda r: sum(1 for x in r if x.get("ocr_warning"))),
        ("Compliance Score %",  lambda r: f"{sum(1 for x in r if x['status']==STATUS_COMPLIANT)/len(r)*100:.1f}%"
                                          if r else "0%"),
    ]

    row_fills = [None, CLR_COMPLIANT, CLR_NON_COMPLY, CLR_NOT_FOUND, CLR_OCR_WARN, None]

    for row_i, (label, fn) in enumerate(metrics, 3):
        ws.cell(row=row_i, column=1, value=label).font = _font(bold=True)
        for v_i, vendor in enumerate(vendors, 2):
            results   = all_results.get(vendor, [])
            val       = fn(results)
            cell      = ws.cell(row=row_i, column=v_i, value=val)
            cell.font = _font(bold=(row_i == 8), size=11 if row_i == 8 else 10)
            cell.alignment = _center()
            if row_fills[row_i - 3]:
                cell.fill = _fill(row_fills[row_i - 3])

    # Rank by compliance score (row 9 onwards)
    ws.cell(row=9, column=1, value="Rank").font = _font(bold=True)
    scores = {}
    for vendor, results in all_results.items():
        if results:
            scores[vendor] = sum(1 for r in results if r["status"] == STATUS_COMPLIANT) / len(results)

    ranked = sorted(scores, key=scores.get, reverse=True)
    for v_i, vendor in enumerate(vendors, 2):
        rank = ranked.index(vendor) + 1 if vendor in ranked else "N/A"
        cell = ws.cell(row=9, column=v_i, value=f"#{rank}")
        cell.font      = _font(bold=True, size=12)
        cell.alignment = _center()
        cell.fill      = _fill(CLR_COMPLIANT if rank == 1 else CLR_SUMMARY_BG)


# =====================================
# LOAD RESULTS
# =====================================
def load_results(vendor: str) -> list[dict]:
    path = RESULTS_DIR / f"{vendor}_results.json"
    if not path.exists():
        print(f"  WARNING: No results file found for vendor '{vendor}' at {path}")
        return []
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def discover_vendor_results() -> list[str]:
    if not RESULTS_DIR.exists():
        return []
    return [
        p.stem.replace("_results", "")
        for p in RESULTS_DIR.glob("*_results.json")
    ]


# =====================================
# MAIN
# =====================================
def main():
    parser = argparse.ArgumentParser(
        description="Generate Excel compliance reports from engine results"
    )
    parser.add_argument(
        "--vendor", type=str, default=None,
        help="Generate report for a specific vendor only."
    )
    args = parser.parse_args()

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    if args.vendor:
        vendors = [args.vendor]
    else:
        vendors = discover_vendor_results()
        if not vendors:
            sys.exit(
                f"ERROR: No result files found in {RESULTS_DIR}.\n"
                f"Run compliance_engine.py first."
            )

    print(f"\nGenerating reports for: {vendors}\n")

    all_results = {}

    for vendor in vendors:
        results = load_results(vendor)
        if not results:
            continue

        all_results[vendor] = results

        # Build + save individual vendor report
        wb       = build_vendor_report(vendor, results)
        out_path = REPORTS_DIR / f"{vendor}_Compliance.xlsx"
        wb.save(str(out_path))
        print(f"  ✓ {out_path.resolve()}")

    # Build comparison report (only if multiple vendors)
    if len(all_results) > 1:
        comp_wb   = build_comparison_report(all_results)
        comp_path = REPORTS_DIR / "Comparison_Summary.xlsx"
        comp_wb.save(str(comp_path))
        print(f"  ✓ {comp_path.resolve()}")

    print(f"\nAll reports saved to: {REPORTS_DIR.resolve()}")


if __name__ == "__main__":
    main()
